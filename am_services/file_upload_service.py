"""
File Upload Service
Handles file uploads, storage, and Excel sheet splitting
"""
import os
import uuid
import shutil
from pathlib import Path
from typing import List, Optional, Tuple
from datetime import datetime
import pandas as pd
from fastapi import UploadFile
import openpyxl

from am_common.upload_models import FileUpload, FileType, ProcessingStatus, SheetInfo


class FileUploadService:
    """Service for handling file uploads and processing"""
    
    def __init__(self, upload_dir: str = "data/uploads", sheets_dir: str = "data/sheets"):
        self.upload_dir = Path(upload_dir)
        self.sheets_dir = Path(sheets_dir)
        
        # Create directories if they don't exist
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        self.sheets_dir.mkdir(parents=True, exist_ok=True)
        
    def generate_unique_id(self) -> str:
        """Generate a unique ID for files"""
        return str(uuid.uuid4())
    
    def get_file_type(self, filename: str) -> FileType:
        """Determine file type from filename"""
        extension = Path(filename).suffix.lower()
        if extension in ['.xlsx', '.xls']:
            return FileType.EXCEL
        elif extension == '.csv':
            return FileType.CSV
        else:
            raise ValueError(f"Unsupported file type: {extension}")
    
    async def save_uploaded_file(self, file: UploadFile) -> FileUpload:
        """Save uploaded file and return FileUpload object"""
        # Generate unique ID and filename
        file_id = self.generate_unique_id()
        file_type = self.get_file_type(file.filename)
        stored_filename = f"{file_id}_{file.filename}"
        file_path = self.upload_dir / stored_filename
        
        # Save file to disk
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Create FileUpload object
        file_upload = FileUpload(
            file_id=file_id,
            original_filename=file.filename,
            stored_filename=stored_filename,
            file_type=file_type,
            file_path=str(file_path),
            file_size=len(content),
            status=ProcessingStatus.UPLOADED
        )
        
        return file_upload
    
    def get_excel_sheet_info(self, file_path: str) -> List[SheetInfo]:
        """Get information about all sheets in an Excel file"""
        try:
            # Use openpyxl to get sheet names and basic info
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets_info = []
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Get dimensions
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                sheet_info = SheetInfo(
                    sheet_name=sheet_name,
                    row_count=max_row if max_row else 0,
                    column_count=max_col if max_col else 0,
                    file_id=""  # Will be set by caller
                )
                sheets_info.append(sheet_info)
            
            workbook.close()
            return sheets_info
            
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
    
    def split_excel_into_sheets(self, parent_file: FileUpload) -> List[FileUpload]:
        """Split Excel file into individual sheet files"""
        if parent_file.file_type != FileType.EXCEL:
            raise ValueError("Can only split Excel files")
        
        sheet_files = []
        
        try:
            # Read Excel file
            excel_data = pd.read_excel(parent_file.file_path, sheet_name=None, engine='openpyxl')
            
            for sheet_name, sheet_df in excel_data.items():
                # Generate unique ID for sheet
                sheet_id = self.generate_unique_id()
                
                # Create filename for individual sheet
                base_name = Path(parent_file.original_filename).stem
                sheet_filename = f"{sheet_id}_{base_name}_{sheet_name}.xlsx"
                sheet_path = self.sheets_dir / sheet_filename
                
                # Save individual sheet as Excel file
                with pd.ExcelWriter(sheet_path, engine='openpyxl') as writer:
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Create FileUpload object for sheet
                sheet_file = FileUpload(
                    file_id=sheet_id,
                    original_filename=f"{base_name}_{sheet_name}.xlsx",
                    stored_filename=sheet_filename,
                    file_type=FileType.SHEET,
                    file_path=str(sheet_path),
                    parent_id=parent_file.file_id,
                    sheet_name=sheet_name,
                    status=ProcessingStatus.UPLOADED,
                    file_size=os.path.getsize(sheet_path) if os.path.exists(sheet_path) else 0
                )
                
                sheet_files.append(sheet_file)
        
        except Exception as e:
            raise ValueError(f"Error splitting Excel file: {str(e)}")
        
        return sheet_files
    
    def update_file_status(self, file_upload: FileUpload, status: ProcessingStatus, 
                          error_message: Optional[str] = None) -> FileUpload:
        """Update file processing status"""
        file_upload.status = status
        file_upload.updated_at = datetime.utcnow()
        if error_message:
            file_upload.error_message = error_message
        return file_upload
    
    def cleanup_file(self, file_path: str) -> bool:
        """Remove file from disk"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                return True
            return False
        except Exception:
            return False
    
    def get_file_info(self, file_id: str, file_uploads: List[FileUpload]) -> Optional[FileUpload]:
        """Get file info by ID"""
        for file_upload in file_uploads:
            if file_upload.file_id == file_id:
                return file_upload
        return None